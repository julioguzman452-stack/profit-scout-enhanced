"""Profit Scout AI — FastAPI backend.

Mock eBay layer + Gemini 3.1 Pro Preview AI + JWT auth + pallet mode.
"""
from __future__ import annotations

import asyncio
import base64
import csv
import io
import json
import logging
import os
import re
import time
import uuid
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, List, Optional
from urllib.parse import quote_plus

import bcrypt
import httpx
import jwt
from dotenv import load_dotenv
from fastapi import APIRouter, Depends, FastAPI, File, HTTPException, UploadFile
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, EmailStr, Field
from starlette.middleware.cors import CORSMiddleware

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
logger = logging.getLogger("profit-scout")

# ---------------- Mongo ----------------
mongo_url = os.environ["MONGO_URL"]
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ["DB_NAME"]]
users_col = db["users"]
history_col = db["history"]
pallets_col = db["pallets"]
pallet_items_col = db["pallet_items"]
fb_comps_col = db["fb_comps"]
inventory_col = db["inventory"]
settings_col = db["settings"]

# ---------------- JWT ----------------
JWT_SECRET = os.environ["JWT_SECRET_KEY"]
JWT_ALGO = os.environ.get("JWT_ALGORITHM", "HS256")
JWT_EXP_MIN = int(os.environ.get("JWT_EXPIRE_MINUTES", "10080"))

security = HTTPBearer(auto_error=False)


def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()


def verify_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode(), hashed.encode())
    except Exception:
        return False


def make_token(user_id: str, email: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "exp": datetime.now(timezone.utc) + timedelta(minutes=JWT_EXP_MIN),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGO)


async def get_current_user(creds: Optional[HTTPAuthorizationCredentials] = Depends(security)) -> dict:
    if not creds or not creds.credentials:
        raise HTTPException(status_code=401, detail="Missing authentication token")
    try:
        payload = jwt.decode(creds.credentials, JWT_SECRET, algorithms=[JWT_ALGO])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid token")
    user_id = payload.get("sub")
    user = await users_col.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    return user


# ---------------- Schemas ----------------
class RegisterIn(BaseModel):
    email: EmailStr
    password: str = Field(min_length=6)


class LoginIn(BaseModel):
    email: EmailStr
    password: str


class AuthOut(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: dict


class IdentifyImageIn(BaseModel):
    image_base64: str


class SearchIn(BaseModel):
    query: str


class ProfitIn(BaseModel):
    item_name: str = "Item"
    buy_cost: float
    sell_price: float
    shipping_cost: float = 0
    ebay_fee_pct: float = 13.25
    tax_cost: float = 0
    extra_cost: float = 0


class HistoryCreate(BaseModel):
    title: str
    query: Optional[str] = None
    image_thumb_b64: Optional[str] = None
    source: str = "manual"  # barcode | camera | manual
    ebay_data: Optional[dict] = None
    ai_insight: Optional[dict] = None
    profit: Optional[dict] = None
    notes: Optional[str] = ""
    tags: List[str] = []


class FbCompCreate(BaseModel):
    item_name: str
    price: float
    location: Optional[str] = ""
    note: Optional[str] = ""


class PalletCreate(BaseModel):
    name: str
    supplier: str = ""
    purchase_date: Optional[str] = None
    purchase_price: float = 0
    shipping_cost: float = 0
    tax_cost: float = 0
    additional_costs: float = 0
    notes: str = ""


class PalletItemCreate(BaseModel):
    product_name: str
    quantity: int = 1
    retail_value: float = 0
    estimated_resale_value: float = 0
    category: str = ""
    status: str = "available"
    sold_price: Optional[float] = None
    notes: str = ""


class PalletItemUpdate(BaseModel):
    product_name: Optional[str] = None
    quantity: Optional[int] = None
    retail_value: Optional[float] = None
    estimated_resale_value: Optional[float] = None
    category: Optional[str] = None
    status: Optional[str] = None
    sold_price: Optional[float] = None
    notes: Optional[str] = None


# ---------------- App ----------------
app = FastAPI(title="Profit Scout AI")
api = APIRouter(prefix="/api")


@api.get("/")
async def root():
    return {"app": "Profit Scout AI", "status": "ok"}


# ---------------- Auth routes ----------------
@api.post("/auth/register", response_model=AuthOut)
async def register(body: RegisterIn):
    email = body.email.lower()
    existing = await users_col.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    uid = str(uuid.uuid4())
    doc = {
        "id": uid,
        "email": email,
        "password_hash": hash_password(body.password),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await users_col.insert_one(doc)
    token = make_token(uid, email)
    return AuthOut(access_token=token, user={"id": uid, "email": email})


@api.post("/auth/login", response_model=AuthOut)
async def login(body: LoginIn):
    email = body.email.lower()
    user = await users_col.find_one({"email": email})
    if not user or not verify_password(body.password, user["password_hash"]):
        raise HTTPException(status_code=400, detail="Invalid email or password")
    token = make_token(user["id"], email)
    return AuthOut(access_token=token, user={"id": user["id"], "email": email})


@api.get("/auth/me")
async def me(user=Depends(get_current_user)):
    return user


# ---------------- eBay Browse API (OAuth client-credentials) ----------------
EBAY_APP_ID = os.environ.get("EBAY_APP_ID", "").strip()
EBAY_CERT_ID = os.environ.get("EBAY_CERT_ID", "").strip()
EBAY_MARKETPLACE_ID = os.environ.get("EBAY_MARKETPLACE_ID", "EBAY_US").strip() or "EBAY_US"
EBAY_OAUTH_URL = "https://api.ebay.com/identity/v1/oauth2/token"
EBAY_BROWSE_URL = "https://api.ebay.com/buy/browse/v1/item_summary/search"

_ebay_token_cache: dict = {"token": None, "expires_at": 0.0}


def ebay_configured() -> bool:
    return bool(EBAY_APP_ID and EBAY_CERT_ID)


async def _ebay_get_token() -> str:
    """Fetch (and cache) an application access token via client-credentials."""
    now = time.time()
    if _ebay_token_cache["token"] and _ebay_token_cache["expires_at"] - 60 > now:
        return _ebay_token_cache["token"]
    if not ebay_configured():
        raise HTTPException(status_code=503, detail="eBay API not configured")
    basic = base64.b64encode(f"{EBAY_APP_ID}:{EBAY_CERT_ID}".encode()).decode()
    headers = {
        "Authorization": f"Basic {basic}",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    body = "grant_type=client_credentials&scope=https%3A%2F%2Fapi.ebay.com%2Foauth%2Fapi_scope"
    async with httpx.AsyncClient(timeout=15.0) as client_http:
        r = await client_http.post(EBAY_OAUTH_URL, content=body, headers=headers)
    if r.status_code != 200:
        logger.error("eBay OAuth failed %s: %s", r.status_code, r.text[:400])
        raise HTTPException(status_code=502, detail="eBay authentication failed")
    j = r.json()
    tok = j.get("access_token")
    exp = int(j.get("expires_in", 7200))
    if not tok:
        raise HTTPException(status_code=502, detail="eBay returned no access token")
    _ebay_token_cache["token"] = tok
    _ebay_token_cache["expires_at"] = now + exp
    return tok


async def ebay_browse_search(query: str, limit: int = 30) -> dict:
    """Call eBay Browse API. Returns real active-listing data only.

    Raises HTTPException 503 if not configured, 502 on API failure.
    """
    token = await _ebay_get_token()
    params = {
        "q": query,
        "limit": str(min(max(limit, 1), 50)),
        "sort": "bestMatch",
    }
    headers = {
        "Authorization": f"Bearer {token}",
        "X-EBAY-C-MARKETPLACE-ID": EBAY_MARKETPLACE_ID,
        "Content-Type": "application/json",
        "Accept": "application/json",
    }
    async with httpx.AsyncClient(timeout=20.0) as client_http:
        r = await client_http.get(EBAY_BROWSE_URL, params=params, headers=headers)
    if r.status_code == 401:
        # token expired mid-flight — clear cache and retry once
        _ebay_token_cache["token"] = None
        _ebay_token_cache["expires_at"] = 0.0
        token = await _ebay_get_token()
        headers["Authorization"] = f"Bearer {token}"
        async with httpx.AsyncClient(timeout=20.0) as client_http:
            r = await client_http.get(EBAY_BROWSE_URL, params=params, headers=headers)
    if r.status_code != 200:
        logger.error("eBay Browse failed %s: %s", r.status_code, r.text[:400])
        raise HTTPException(status_code=502, detail="eBay search failed")
    payload = r.json()
    total = int(payload.get("total") or 0)
    items_raw = payload.get("itemSummaries") or []
    prices: List[float] = []
    listings: List[dict] = []
    for it in items_raw:
        price_obj = it.get("price") or {}
        try:
            val = float(price_obj.get("value", 0) or 0)
        except (TypeError, ValueError):
            val = 0.0
        if val <= 0:
            continue
        prices.append(val)
        ship_val = 0.0
        opts = it.get("shippingOptions") or []
        if opts:
            try:
                ship_val = float(((opts[0] or {}).get("shippingCost") or {}).get("value", 0) or 0)
            except (TypeError, ValueError):
                ship_val = 0.0
        listings.append({
            "title": (it.get("title") or "")[:200],
            "price": round(val, 2),
            "currency": price_obj.get("currency", "USD"),
            "shipping": round(ship_val, 2),
            "condition": it.get("condition") or "",
            "seller": ((it.get("seller") or {}).get("username")) or "",
            "url": it.get("itemWebUrl") or "",
            "image": ((it.get("image") or {}).get("imageUrl")) or "",
        })
    if not prices:
        return {
            "available": True,
            "query": query,
            "active_count": total,
            "avg_price": 0.0,
            "median_price": 0.0,
            "lowest_price": 0.0,
            "highest_price": 0.0,
            "listings": [],
            "data_source": "ebay_browse_api",
            "marketplace": EBAY_MARKETPLACE_ID,
            "message": "No active listings matched this query.",
        }
    prices_sorted = sorted(prices)
    avg = round(sum(prices) / len(prices), 2)
    lo = round(prices_sorted[0], 2)
    hi = round(prices_sorted[-1], 2)
    med = round(prices_sorted[len(prices_sorted) // 2], 2)
    return {
        "available": True,
        "query": query,
        "active_count": total,
        "sample_count": len(prices),
        "avg_price": avg,
        "median_price": med,
        "lowest_price": lo,
        "highest_price": hi,
        "listings": listings[:12],
        "data_source": "ebay_browse_api",
        "marketplace": EBAY_MARKETPLACE_ID,
    }


# ---------------- UPCitemDB free trial lookup ----------------
UPCDB_URL = "https://api.upcitemdb.com/prod/trial/lookup"


async def upc_lookup(code: str) -> Optional[dict]:
    """Look up UPC/EAN via UPCitemDB free trial (no key required, 100/day/IP).

    Returns None if not found or the service is unavailable. Never fabricates.
    """
    if not code or not code.strip():
        return None
    try:
        async with httpx.AsyncClient(timeout=12.0) as client_http:
            r = await client_http.get(UPCDB_URL, params={"upc": code.strip()})
    except Exception:
        logger.exception("UPCitemDB request failed")
        return None
    if r.status_code != 200:
        return None
    try:
        j = r.json()
    except Exception:
        return None
    items = j.get("items") or []
    if not items:
        return None
    it = items[0]
    title = (it.get("title") or "").strip()
    if not title:
        return None
    return {
        "title": title,
        "brand": (it.get("brand") or "").strip(),
        "model": (it.get("model") or "").strip(),
        "category": (it.get("category") or "").strip(),
        "description": (it.get("description") or "").strip()[:500],
        "image": (it.get("images") or [None])[0],
    }


# ---------------- Marketplace deep-link helpers ----------------
def marketplace_links(query: str) -> dict:
    q = quote_plus(query.strip())
    return {
        "ebay": f"https://www.ebay.com/sch/i.html?_nkw={q}&LH_Sold=1&LH_Complete=1",
        "ebay_active": f"https://www.ebay.com/sch/i.html?_nkw={q}",
        "amazon": f"https://www.amazon.com/s?k={q}",
        "facebook": f"https://www.facebook.com/marketplace/search/?query={q}",
        "mercari": f"https://www.mercari.com/search/?keyword={q}",
        "whatnot": f"https://www.whatnot.com/search/{q}",
    }


# ---------------- Emergent AI helpers ----------------
AI_TIMEOUT_SECONDS = 25.0


async def ai_chat(system: str, user_text: str, image_b64: Optional[str] = None) -> str:
    """Call Gemini via emergentintegrations with a hard timeout.

    Raises HTTPException 503 on any AI failure (missing key, import error,
    timeout, upstream error). Never hangs.
    """
    api_key = os.environ.get("EMERGENT_LLM_KEY")
    if not api_key:
        raise HTTPException(status_code=503, detail="AI service unavailable: missing API key")
    try:
        from emergentintegrations.llm.chat import (
            ImageContent,
            LlmChat,
            UserMessage,
        )
    except Exception as e:
        logger.exception("emergentintegrations import failed")
        raise HTTPException(status_code=503, detail=f"AI service unavailable: {e}")

    chat = LlmChat(
        api_key=api_key,
        session_id=f"ps-{uuid.uuid4()}",
        system_message=system,
    ).with_model("gemini", "gemini-2.5-pro")

    file_contents = []
    if image_b64:
        file_contents.append(ImageContent(image_base64=image_b64))

    msg = UserMessage(text=user_text, file_contents=file_contents or None)
    try:
        reply = await asyncio.wait_for(chat.send_message(msg), timeout=AI_TIMEOUT_SECONDS)
    except asyncio.TimeoutError:
        logger.warning("AI call timed out after %ss", AI_TIMEOUT_SECONDS)
        raise HTTPException(status_code=504, detail="AI service timed out. Please try again.")
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("AI call failed")
        raise HTTPException(status_code=503, detail=f"AI service unavailable: {e}")
    return reply if isinstance(reply, str) else str(reply)


def _extract_json(text: str) -> dict:
    """Extract JSON from a possibly-fenced LLM reply."""
    if not text:
        return {}
    # try fenced block
    m = re.search(r"```(?:json)?\s*(\{.*?\}|\[.*?\])\s*```", text, re.S)
    raw = m.group(1) if m else text
    # find first JSON object / array
    if not raw.strip().startswith(("{", "[")):
        m2 = re.search(r"(\{.*\}|\[.*\])", raw, re.S)
        if m2:
            raw = m2.group(1)
    try:
        return json.loads(raw)
    except Exception:
        return {}


# ---------------- Scan / Search ----------------
@api.post("/scan/identify-image")
async def identify_image(body: IdentifyImageIn, user=Depends(get_current_user)):
    """Gemini vision: identify product from photo, return structured info."""
    if not body.image_base64 or len(body.image_base64) < 200:
        raise HTTPException(status_code=400, detail="Invalid image payload")
    if len(body.image_base64) > 10 * 1024 * 1024:  # ~10 MB base64 cap
        raise HTTPException(status_code=413, detail="Image too large. Please retake at lower quality.")
    system = (
        "You are a reseller product identification expert. From a single product photo "
        "extract a clean JSON object with these keys exactly: product_name, brand, model, "
        "category, condition, ebay_search_keywords (array of 3-6 strings), confidence (0-1). "
        "If the photo does not clearly show a product, return "
        '{"product_name":"Unknown item","confidence":0.0,"ebay_search_keywords":[]}. '
        "Return ONLY raw JSON, no commentary, no markdown."
    )
    reply = await ai_chat(system, "Identify this item for eBay resale.", image_b64=body.image_base64)
    data = _extract_json(reply)
    if not data:
        return {
            "product_name": "Unknown item",
            "ebay_search_keywords": [],
            "confidence": 0.0,
            "error": "Unable to identify product from image.",
        }
    return data


@api.post("/scan/identify-barcode")
async def identify_barcode(payload: dict, user=Depends(get_current_user)):
    """Real UPC/EAN lookup via UPCitemDB. Never fabricates a product name."""
    code = str(payload.get("barcode", "")).strip()
    btype = str(payload.get("type", "")).strip()
    if not code:
        raise HTTPException(status_code=400, detail="barcode required")
    product = await upc_lookup(code)
    if product:
        return {
            "barcode": code,
            "barcode_type": btype or "unknown",
            "lookup_status": "found",
            "product": product,
            "data_source": "upcitemdb",
            "search_query": product["title"],
        }
    return {
        "barcode": code,
        "barcode_type": btype or "unknown",
        "lookup_status": "not_found",
        "lookup_message": "Product not found",
        "data_source": "upcitemdb",
        "search_query": code,
    }


async def _ai_insight_for_ebay(query: str, ebay: dict, buy_cost: float = 0.0, ebay_fee_pct: float = 13.25) -> dict:
    """Generate real-data-only AI insight from live eBay Browse results.

    Never fabricates prices. If AI is unavailable, returns a deterministic
    fallback derived from the real numbers.
    """
    avg = float(ebay.get("avg_price") or 0)
    lo = float(ebay.get("lowest_price") or 0)
    hi = float(ebay.get("highest_price") or 0)
    med = float(ebay.get("median_price") or avg)
    active = int(ebay.get("active_count") or 0)
    listings_preview = ebay.get("listings") or []

    # Deterministic profit math on the REAL median price
    fee_rate = ebay_fee_pct / 100.0
    expected_sale = med if med > 0 else avg
    fees = round(expected_sale * fee_rate, 2)
    expected_profit = round(expected_sale - fees - buy_cost, 2) if expected_sale > 0 else 0.0
    roi_pct = round((expected_profit / buy_cost) * 100, 1) if buy_cost > 0 else 0.0

    system = (
        "You are a reseller analyst. You will receive REAL live eBay Browse data. "
        "Analyze the numbers to produce a decision. NEVER invent prices, listing counts, "
        "or sold data. Only reason about the numbers provided. "
        "Return ONLY raw JSON with these keys: "
        '{"verdict": "BUY"|"MAYBE BUY"|"AVOID", '
        '"risk_level": "Low"|"Medium"|"High", '
        '"sell_through_recommendation": string (1 sentence, plain english), '
        '"reasoning": string (2 short sentences citing the real numbers)}. '
        "No markdown, no extra keys."
    )
    prompt_payload = {
        "query": query,
        "buy_cost": buy_cost,
        "ebay_fee_pct": ebay_fee_pct,
        "computed_expected_profit": expected_profit,
        "computed_roi_pct": roi_pct,
        "ebay_active_count": active,
        "ebay_avg_price": avg,
        "ebay_median_price": med,
        "ebay_low_price": lo,
        "ebay_high_price": hi,
        "sample_titles": [it.get("title", "")[:80] for it in listings_preview[:5]],
    }
    ai: dict = {}
    try:
        reply = await ai_chat(system, json.dumps(prompt_payload))
        ai = _extract_json(reply) or {}
    except HTTPException as e:
        logger.info("AI insight fallback: %s", e.detail)
        ai = {}

    verdict = str(ai.get("verdict") or "").upper()
    if verdict not in {"BUY", "MAYBE BUY", "AVOID"}:
        # deterministic fallback verdict
        if buy_cost > 0 and expected_profit >= max(5, buy_cost * 0.5):
            verdict = "BUY"
        elif buy_cost > 0 and expected_profit >= max(2, buy_cost * 0.2):
            verdict = "MAYBE BUY"
        elif buy_cost <= 0 and expected_sale > 0:
            verdict = "MAYBE BUY"
        else:
            verdict = "AVOID"

    risk_level = str(ai.get("risk_level") or "").capitalize()
    if risk_level not in {"Low", "Medium", "High"}:
        if active <= 5:
            risk_level = "High"
        elif active >= 50:
            risk_level = "Low"
        else:
            risk_level = "Medium"

    sell_rec = str(ai.get("sell_through_recommendation") or "").strip()
    if not sell_rec:
        if active >= 100:
            sell_rec = f"Highly competitive category with {active} active listings — price competitively and use fast shipping."
        elif active >= 20:
            sell_rec = f"Moderate competition ({active} active listings). Aim near median price for a reasonable sell-through window."
        elif active >= 1:
            sell_rec = f"Low competition ({active} active listings) — you can price near the higher end."
        else:
            sell_rec = "No active competition found; niche demand — sell-through may be slow."

    reasoning = str(ai.get("reasoning") or "").strip()
    if not reasoning:
        reasoning = (
            f"Median active eBay price is ${med:.2f} across {active} listings "
            f"(range ${lo:.2f}–${hi:.2f}). "
            + (f"Expected profit at that price is ${expected_profit:.2f} ({roi_pct}% ROI)." if buy_cost > 0
               else "Provide a buy cost to compute exact ROI.")
        )

    return {
        "verdict": verdict,
        "risk_level": risk_level,
        "sell_through_recommendation": sell_rec,
        "reasoning": reasoning,
        "expected_sale_price": round(expected_sale, 2),
        "estimated_low": round(lo, 2),
        "estimated_high": round(hi, 2),
        "expected_profit": expected_profit,
        "roi_pct": roi_pct,
        "ebay_fee_estimated": fees,
        "based_on": "ebay_browse_api",
        "sample_size": ebay.get("sample_count") or len(listings_preview),
    }


@api.post("/search")
async def search(body: SearchIn, user=Depends(get_current_user)):
    """Live eBay Browse search + AI insight on the real data. No mock data.

    Response shape:
      {
        "query": str,
        "ebay": {...} | null,
        "pricing_available": bool,
        "pricing_message": str,   # user-facing reason if not available
        "ai_insight": {...} | null,
        "marketplace_links": {ebay, ebay_active, amazon, facebook, mercari, whatnot}
      }
    """
    query = body.query.strip()
    if not query:
        raise HTTPException(status_code=400, detail="query required")

    links = marketplace_links(query)
    if not ebay_configured():
        return {
            "query": query,
            "ebay": None,
            "pricing_available": False,
            "pricing_message": "eBay API not configured",
            "ai_insight": None,
            "marketplace_links": links,
        }

    try:
        ebay = await ebay_browse_search(query)
    except HTTPException as e:
        logger.info("eBay live fetch failed: %s", e.detail)
        return {
            "query": query,
            "ebay": None,
            "pricing_available": False,
            "pricing_message": "Live pricing unavailable" if e.status_code != 503 else "eBay API not configured",
            "ai_insight": None,
            "marketplace_links": links,
        }

    if not ebay.get("listings"):
        return {
            "query": query,
            "ebay": ebay,
            "pricing_available": False,
            "pricing_message": "No live eBay listings found for this query.",
            "ai_insight": None,
            "marketplace_links": links,
        }

    insight = await _ai_insight_for_ebay(query, ebay)
    return {
        "query": query,
        "ebay": ebay,
        "pricing_available": True,
        "pricing_message": "",
        "ai_insight": insight,
        "marketplace_links": links,
    }


@api.post("/profit/verdict")
async def profit_verdict(body: ProfitIn, user=Depends(get_current_user)):
    fees = body.sell_price * (body.ebay_fee_pct / 100.0)
    total_cost = body.buy_cost + body.shipping_cost + body.tax_cost + body.extra_cost + fees
    net = round(body.sell_price - total_cost, 2)
    roi = round((net / body.buy_cost) * 100, 1) if body.buy_cost > 0 else 0.0
    fixed_costs = body.buy_cost + body.shipping_cost + body.tax_cost + body.extra_cost
    fee_rate = body.ebay_fee_pct / 100.0
    # break-even sale price: sale_price - (fixed + sale_price*fee_rate) = 0 -> sale_price = fixed / (1 - fee_rate)
    break_even_price = round(fixed_costs / (1 - fee_rate), 2) if fee_rate < 1 else 0.0
    margin_pct = round((net / body.sell_price) * 100, 1) if body.sell_price > 0 else 0.0

    if roi >= 50 and net >= 5:
        verdict = "BUY"
    elif roi >= 20 and net >= 2:
        verdict = "MAYBE BUY"
    else:
        verdict = "AVOID"

    # AI explanation (best-effort)
    explanation = ""
    try:
        system = (
            "You are a reseller mentor. In 2 short plain-English sentences, explain why this deal "
            "is good/risky/bad for a hobby flipper. No markdown."
        )
        prompt = (
            f"Item: {body.item_name}. Buy ${body.buy_cost}. Sell ${body.sell_price}. "
            f"Fees ${round(fees,2)}. Shipping ${body.shipping_cost}. Tax ${body.tax_cost}. "
            f"Net ${net}. ROI {roi}%. Verdict: {verdict}."
        )
        explanation = (await ai_chat(system, prompt)).strip()
    except HTTPException:
        explanation = f"{verdict}: net profit ${net} ({roi}% ROI)."

    return {
        "verdict": verdict,
        "net_profit": net,
        "roi_pct": roi,
        "profit_margin_pct": margin_pct,
        "break_even_price": break_even_price,
        "ebay_fee": round(fees, 2),
        "total_cost": round(total_cost, 2),
        "explanation": explanation,
    }


# ---------------- History ----------------
@api.post("/history")
async def create_history(body: HistoryCreate, user=Depends(get_current_user)):
    doc = body.dict()
    doc["id"] = str(uuid.uuid4())
    doc["user_id"] = user["id"]
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await history_col.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.get("/history")
async def list_history(user=Depends(get_current_user)):
    cur = history_col.find({"user_id": user["id"]}, {"_id": 0}).sort("created_at", -1).limit(200)
    return await cur.to_list(length=200)


@api.delete("/history/{hid}")
async def delete_history(hid: str, user=Depends(get_current_user)):
    res = await history_col.delete_one({"id": hid, "user_id": user["id"]})
    if not res.deleted_count:
        raise HTTPException(status_code=404, detail="not found")
    return {"ok": True}


# ---------------- FB Marketplace comps ----------------
@api.post("/fb-comps")
async def create_fb_comp(body: FbCompCreate, user=Depends(get_current_user)):
    doc = body.dict()
    doc["id"] = str(uuid.uuid4())
    doc["user_id"] = user["id"]
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await fb_comps_col.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.get("/fb-comps")
async def list_fb_comps(item: Optional[str] = None, user=Depends(get_current_user)):
    q: dict = {"user_id": user["id"]}
    if item:
        q["item_name"] = {"$regex": re.escape(item), "$options": "i"}
    cur = fb_comps_col.find(q, {"_id": 0}).sort("created_at", -1).limit(200)
    return await cur.to_list(length=200)


@api.delete("/fb-comps/{cid}")
async def delete_fb_comp(cid: str, user=Depends(get_current_user)):
    res = await fb_comps_col.delete_one({"id": cid, "user_id": user["id"]})
    if not res.deleted_count:
        raise HTTPException(status_code=404, detail="not found")
    return {"ok": True}


# ---------------- Pallets ----------------
def _pallet_total_investment(p: dict) -> float:
    return round(
        float(p.get("purchase_price", 0) or 0)
        + float(p.get("shipping_cost", 0) or 0)
        + float(p.get("tax_cost", 0) or 0)
        + float(p.get("additional_costs", 0) or 0),
        2,
    )


@api.post("/pallets")
async def create_pallet(body: PalletCreate, user=Depends(get_current_user)):
    doc = body.dict()
    doc["id"] = str(uuid.uuid4())
    doc["user_id"] = user["id"]
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    doc["total_investment"] = _pallet_total_investment(doc)
    await pallets_col.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.get("/pallets")
async def list_pallets(user=Depends(get_current_user)):
    cur = pallets_col.find({"user_id": user["id"]}, {"_id": 0}).sort("created_at", -1)
    pallets = await cur.to_list(length=500)
    if not pallets:
        return []
    pids = [p["id"] for p in pallets]
    # Single batch fetch of all items across these pallets
    items_cur = pallet_items_col.find(
        {"pallet_id": {"$in": pids}, "user_id": user["id"]}, {"_id": 0}
    )
    all_items = await items_cur.to_list(length=20000)
    items_by_pallet: dict = {}
    for it in all_items:
        items_by_pallet.setdefault(it["pallet_id"], []).append(it)
    for p in pallets:
        p["dashboard"] = _compute_dashboard_from_items(p, items_by_pallet.get(p["id"], []))
    return pallets


def _compute_dashboard_from_items(pallet: dict, items: list) -> dict:
    """Same math as _compute_dashboard but takes pre-fetched items."""
    total_investment = _pallet_total_investment(pallet)
    revenue = 0.0
    remaining_value = 0.0
    counts = {"available": 0, "listed": 0, "sold": 0, "damaged": 0, "returned": 0, "missing": 0}
    for it in items:
        qty = int(it.get("quantity", 1) or 1)
        st = (it.get("status") or "available").lower()
        if st not in counts:
            st = "available"
        counts[st] += qty
        if st == "sold":
            sp = it.get("sold_price")
            if sp is None:
                sp = it.get("estimated_resale_value") or 0
            revenue += float(sp) * qty
        elif st in ("available", "listed"):
            remaining_value += float(it.get("estimated_resale_value") or 0) * qty
    revenue = round(revenue, 2)
    remaining_value = round(remaining_value, 2)
    profit = round(revenue - total_investment, 2)
    break_even_pct = round((revenue / total_investment) * 100, 1) if total_investment > 0 else 0.0
    break_even_remaining = round(max(total_investment - revenue, 0), 2)
    estimated_final = round(revenue + remaining_value - total_investment, 2)
    total_items = sum(counts.values())
    return {
        "total_investment": total_investment,
        "revenue_recovered": revenue,
        "current_profit": profit,
        "break_even_percent": break_even_pct,
        "break_even_remaining": break_even_remaining,
        "remaining_inventory_value": remaining_value,
        "estimated_final_profit": estimated_final,
        "counts": counts,
        "total_items": total_items,
    }


@api.get("/pallets/{pid}")
async def get_pallet(pid: str, user=Depends(get_current_user)):
    p = await pallets_col.find_one({"id": pid, "user_id": user["id"]}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Pallet not found")
    p["dashboard"] = await _compute_dashboard(p)
    return p


@api.delete("/pallets/{pid}")
async def delete_pallet(pid: str, user=Depends(get_current_user)):
    res = await pallets_col.delete_one({"id": pid, "user_id": user["id"]})
    await pallet_items_col.delete_many({"pallet_id": pid, "user_id": user["id"]})
    if not res.deleted_count:
        raise HTTPException(status_code=404, detail="not found")
    return {"ok": True}


@api.post("/pallets/{pid}/items")
async def add_pallet_item(pid: str, body: PalletItemCreate, user=Depends(get_current_user)):
    p = await pallets_col.find_one({"id": pid, "user_id": user["id"]})
    if not p:
        raise HTTPException(status_code=404, detail="Pallet not found")
    doc = body.dict()
    doc["id"] = str(uuid.uuid4())
    doc["pallet_id"] = pid
    doc["user_id"] = user["id"]
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await pallet_items_col.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.get("/pallets/{pid}/items")
async def list_pallet_items(pid: str, user=Depends(get_current_user)):
    cur = pallet_items_col.find({"pallet_id": pid, "user_id": user["id"]}, {"_id": 0}).sort("created_at", 1)
    return await cur.to_list(length=5000)


@api.patch("/pallets/{pid}/items/{iid}")
async def update_pallet_item(pid: str, iid: str, body: PalletItemUpdate, user=Depends(get_current_user)):
    updates = {k: v for k, v in body.dict().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")
    res = await pallet_items_col.update_one(
        {"id": iid, "pallet_id": pid, "user_id": user["id"]}, {"$set": updates}
    )
    if not res.matched_count:
        raise HTTPException(status_code=404, detail="Item not found")
    item = await pallet_items_col.find_one({"id": iid}, {"_id": 0})
    return item


@api.delete("/pallets/{pid}/items/{iid}")
async def delete_pallet_item(pid: str, iid: str, user=Depends(get_current_user)):
    res = await pallet_items_col.delete_one({"id": iid, "pallet_id": pid, "user_id": user["id"]})
    if not res.deleted_count:
        raise HTTPException(status_code=404, detail="not found")
    return {"ok": True}


# Compute pallet dashboard stats
async def _compute_dashboard(pallet: dict) -> dict:
    items = await pallet_items_col.find(
        {"pallet_id": pallet["id"], "user_id": pallet["user_id"]}, {"_id": 0}
    ).to_list(length=10000)
    total_investment = _pallet_total_investment(pallet)
    revenue = 0.0
    remaining_value = 0.0
    counts = {"available": 0, "listed": 0, "sold": 0, "damaged": 0, "returned": 0, "missing": 0}
    for it in items:
        qty = int(it.get("quantity", 1) or 1)
        st = (it.get("status") or "available").lower()
        if st not in counts:
            st = "available"
        counts[st] += qty
        if st == "sold":
            sp = it.get("sold_price")
            if sp is None:
                sp = it.get("estimated_resale_value") or 0
            revenue += float(sp) * qty
        elif st in ("available", "listed"):
            remaining_value += float(it.get("estimated_resale_value") or 0) * qty
    revenue = round(revenue, 2)
    remaining_value = round(remaining_value, 2)
    profit = round(revenue - total_investment, 2)
    break_even_pct = round((revenue / total_investment) * 100, 1) if total_investment > 0 else 0.0
    break_even_remaining = round(max(total_investment - revenue, 0), 2)
    estimated_final = round(revenue + remaining_value - total_investment, 2)
    total_items = sum(counts.values())
    return {
        "total_investment": total_investment,
        "revenue_recovered": revenue,
        "current_profit": profit,
        "break_even_percent": break_even_pct,
        "break_even_remaining": break_even_remaining,
        "remaining_inventory_value": remaining_value,
        "estimated_final_profit": estimated_final,
        "counts": counts,
        "total_items": total_items,
    }


# ---------------- Manifest import ----------------
def _parse_manifest_bytes(filename: str, data: bytes) -> str:
    """Convert uploaded manifest to plain text for AI extraction."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        try:
            text = data.decode("utf-8", errors="ignore")
        except Exception:
            text = data.decode("latin-1", errors="ignore")
        return text[:20000]
    if name.endswith(".xlsx") or name.endswith(".xls"):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
            out = []
            for ws in wb.worksheets:
                out.append(f"# Sheet: {ws.title}")
                for row in ws.iter_rows(values_only=True):
                    cells = ["" if c is None else str(c) for c in row]
                    out.append(",".join(cells))
            return "\n".join(out)[:20000]
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"xlsx parse failed: {e}")
    if name.endswith(".pdf"):
        try:
            from pypdf import PdfReader
            reader = PdfReader(io.BytesIO(data))
            text = "\n".join((p.extract_text() or "") for p in reader.pages)
            return text[:20000]
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"pdf parse failed: {e}")
    # fallback: treat as text
    try:
        return data.decode("utf-8", errors="ignore")[:20000]
    except Exception:
        raise HTTPException(status_code=400, detail="Unsupported file format")


@api.post("/pallets/{pid}/manifest")
async def upload_manifest(
    pid: str,
    file: UploadFile = File(...),
    user=Depends(get_current_user),
):
    p = await pallets_col.find_one({"id": pid, "user_id": user["id"]})
    if not p:
        raise HTTPException(status_code=404, detail="Pallet not found")
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Empty file")
    text = _parse_manifest_bytes(file.filename or "", raw)

    system = (
        "You are a liquidation pallet manifest parser. From the raw text below, extract every line item. "
        'Return ONLY raw JSON with shape: {"items": [{"product_name": str, "quantity": int, '
        '"retail_value": number, "estimated_resale_value": number, "category": str}]}. '
        "If retail or resale value is missing, estimate retail conservatively and set resale to 35-50% of retail. "
        "Return AT MOST 60 items, prioritizing higher-value items. No commentary, no markdown."
    )
    reply = await ai_chat(system, f"Manifest text:\n{text}")
    data = _extract_json(reply)
    items = data.get("items") if isinstance(data, dict) else None
    if not isinstance(items, list):
        items = []

    inserted = []
    now = datetime.now(timezone.utc).isoformat()
    for it in items[:60]:
        try:
            qty = int(it.get("quantity") or 1)
        except Exception:
            qty = 1
        doc = {
            "id": str(uuid.uuid4()),
            "pallet_id": pid,
            "user_id": user["id"],
            "product_name": str(it.get("product_name") or "Unknown")[:200],
            "quantity": max(qty, 1),
            "retail_value": float(it.get("retail_value") or 0) or 0.0,
            "estimated_resale_value": float(it.get("estimated_resale_value") or 0) or 0.0,
            "category": str(it.get("category") or "")[:80],
            "status": "available",
            "sold_price": None,
            "notes": "",
            "created_at": now,
        }
        inserted.append(doc)
    if inserted:
        await pallet_items_col.insert_many(inserted)
    return {"imported": len(inserted)}


@api.get("/pallets/{pid}/analysis")
async def pallet_analysis(pid: str, user=Depends(get_current_user)):
    p = await pallets_col.find_one({"id": pid, "user_id": user["id"]}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Pallet not found")
    items = await pallet_items_col.find(
        {"pallet_id": pid, "user_id": user["id"]}, {"_id": 0}
    ).to_list(length=2000)
    dashboard = await _compute_dashboard(p)

    # Quick top-by-value (deterministic)
    by_value = sorted(items, key=lambda x: float(x.get("estimated_resale_value", 0) or 0) * int(x.get("quantity", 1) or 1), reverse=True)
    top_value = [{"name": i["product_name"], "value": round(float(i.get("estimated_resale_value", 0) or 0) * int(i.get("quantity", 1) or 1), 2)} for i in by_value[:10]]

    # AI analysis (best-effort)
    ai: dict = {}
    try:
        # Compact item summary
        compact = [
            {"n": i["product_name"][:60], "q": i.get("quantity", 1), "r": i.get("estimated_resale_value", 0), "c": i.get("category", "")}
            for i in items[:60]
        ]
        system = (
            "You are a pallet liquidation analyst. Given the pallet items JSON, return ONLY raw JSON with keys: "
            "fastest_moving (array of up to 10 product names — items most likely to sell quickly), "
            "high_risk (array of up to 10 product names that may be hard to sell), "
            "slow_moving (array of up to 10 product names likely to take a long time), "
            "recommended_listing_order (array of up to 15 product names in best listing order), "
            "forecast (object with conservative, expected, best_case numbers — total revenue estimate). "
            "No markdown."
        )
        prompt = json.dumps({"investment": dashboard["total_investment"], "items": compact})
        reply = await ai_chat(system, prompt)
        ai = _extract_json(reply) or {}
    except HTTPException:
        ai = {}

    # Always include deterministic top_value; fill in defaults if AI failed
    return {
        "dashboard": dashboard,
        "top_value_items": top_value,
        "fastest_moving": ai.get("fastest_moving", []),
        "high_risk": ai.get("high_risk", []),
        "slow_moving": ai.get("slow_moving", []),
        "recommended_listing_order": ai.get("recommended_listing_order", []),
        "forecast": ai.get("forecast", {
            "conservative": round(dashboard["remaining_inventory_value"] * 0.6 + dashboard["revenue_recovered"], 2),
            "expected": round(dashboard["remaining_inventory_value"] * 0.85 + dashboard["revenue_recovered"], 2),
            "best_case": round(dashboard["remaining_inventory_value"] * 1.1 + dashboard["revenue_recovered"], 2),
        }),
    }


# ---------------- Inventory & Sales ----------------
SOURCE_TAGS = [
    "Whatnot", "Pallet", "Facebook", "Garage Sale", "Flea Market",
    "Auction", "Thrift Store", "Retail Arbitrage", "Other",
]
SALE_PLATFORMS = ["eBay", "Whatnot", "Facebook Marketplace", "Mercari", "Local Sale"]
STATUSES = ["in_stock", "listed", "sold", "returned"]


class InventoryCreate(BaseModel):
    title: str
    sku: Optional[str] = ""
    category: Optional[str] = ""
    source: str = "Other"
    pallet_id: Optional[str] = None
    purchase_price: float = 0
    sale_price: Optional[float] = None
    fees: float = 0
    shipping: float = 0
    tax: float = 0
    packaging: float = 0
    misc: float = 0
    platform: Optional[str] = None
    status: str = "in_stock"
    date_purchased: Optional[str] = None
    date_listed: Optional[str] = None
    date_sold: Optional[str] = None
    notes: Optional[str] = ""


class InventoryUpdate(BaseModel):
    title: Optional[str] = None
    sku: Optional[str] = None
    category: Optional[str] = None
    source: Optional[str] = None
    pallet_id: Optional[str] = None
    purchase_price: Optional[float] = None
    sale_price: Optional[float] = None
    fees: Optional[float] = None
    shipping: Optional[float] = None
    tax: Optional[float] = None
    packaging: Optional[float] = None
    misc: Optional[float] = None
    platform: Optional[str] = None
    status: Optional[str] = None
    date_purchased: Optional[str] = None
    date_listed: Optional[str] = None
    date_sold: Optional[str] = None
    notes: Optional[str] = None


def _compute_inventory_metrics(doc: dict) -> dict:
    purchase = float(doc.get("purchase_price", 0) or 0)
    sale = doc.get("sale_price")
    sale_f = float(sale) if sale is not None else 0.0
    fees = float(doc.get("fees", 0) or 0)
    shipping = float(doc.get("shipping", 0) or 0)
    tax = float(doc.get("tax", 0) or 0)
    packaging = float(doc.get("packaging", 0) or 0)
    misc = float(doc.get("misc", 0) or 0)
    total_cost = round(purchase + fees + shipping + tax + packaging + misc, 2)
    net = round(sale_f - total_cost, 2) if sale is not None else 0.0
    roi = round((net / purchase) * 100, 1) if purchase > 0 and sale is not None else 0.0
    margin = round((net / sale_f) * 100, 1) if sale_f > 0 else 0.0
    days_to_sell = None
    try:
        if doc.get("date_purchased") and doc.get("date_sold"):
            d1 = datetime.fromisoformat(doc["date_purchased"].replace("Z", "+00:00"))
            d2 = datetime.fromisoformat(doc["date_sold"].replace("Z", "+00:00"))
            days_to_sell = max(0, (d2 - d1).days)
    except Exception:
        days_to_sell = None
    return {
        "total_cost": total_cost,
        "net_profit": net,
        "roi_pct": roi,
        "profit_margin_pct": margin,
        "days_to_sell": days_to_sell,
    }


@api.post("/inventory")
async def create_inventory(body: InventoryCreate, user=Depends(get_current_user)):
    doc = body.dict()
    doc["id"] = str(uuid.uuid4())
    doc["user_id"] = user["id"]
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    if not doc.get("date_purchased"):
        doc["date_purchased"] = doc["created_at"]
    await inventory_col.insert_one(doc)
    doc.pop("_id", None)
    doc["metrics"] = _compute_inventory_metrics(doc)
    return doc


@api.get("/inventory")
async def list_inventory(
    status: Optional[str] = None,
    source: Optional[str] = None,
    user=Depends(get_current_user),
):
    q: dict = {"user_id": user["id"]}
    if status:
        q["status"] = status
    if source:
        q["source"] = source
    cur = inventory_col.find(q, {"_id": 0}).sort("created_at", -1).limit(2000)
    items = await cur.to_list(length=2000)
    for it in items:
        it["metrics"] = _compute_inventory_metrics(it)
    return items


@api.get("/inventory/export")
async def export_inventory(user=Depends(get_current_user)):
    items = await inventory_col.find({"user_id": user["id"]}, {"_id": 0}).to_list(length=5000)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "id", "title", "sku", "category", "source", "platform", "status",
        "purchase_price", "sale_price", "fees", "shipping", "tax", "packaging", "misc",
        "net_profit", "roi_pct", "profit_margin_pct",
        "date_purchased", "date_listed", "date_sold", "notes",
    ])
    for it in items:
        m = _compute_inventory_metrics(it)
        writer.writerow([
            it.get("id", ""), it.get("title", ""), it.get("sku", ""),
            it.get("category", ""), it.get("source", ""),
            it.get("platform", "") or "", it.get("status", ""),
            it.get("purchase_price", 0), it.get("sale_price") or "",
            it.get("fees", 0), it.get("shipping", 0), it.get("tax", 0),
            it.get("packaging", 0), it.get("misc", 0),
            m["net_profit"], m["roi_pct"], m["profit_margin_pct"],
            it.get("date_purchased", "") or "", it.get("date_listed", "") or "",
            it.get("date_sold", "") or "", (it.get("notes", "") or "").replace("\n", " "),
        ])
    return {"filename": "inventory.csv", "csv": output.getvalue(), "count": len(items)}


@api.get("/inventory/{iid}")
async def get_inventory(iid: str, user=Depends(get_current_user)):
    doc = await inventory_col.find_one({"id": iid, "user_id": user["id"]}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Not found")
    doc["metrics"] = _compute_inventory_metrics(doc)
    return doc


@api.patch("/inventory/{iid}")
async def update_inventory(iid: str, body: InventoryUpdate, user=Depends(get_current_user)):
    updates = {k: v for k, v in body.dict().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")
    # Auto-set date_sold if marking sold without one
    if updates.get("status") == "sold" and not updates.get("date_sold"):
        existing = await inventory_col.find_one({"id": iid, "user_id": user["id"]})
        if existing and not existing.get("date_sold"):
            updates["date_sold"] = datetime.now(timezone.utc).isoformat()
    res = await inventory_col.update_one(
        {"id": iid, "user_id": user["id"]}, {"$set": updates}
    )
    if not res.matched_count:
        raise HTTPException(status_code=404, detail="Not found")
    doc = await inventory_col.find_one({"id": iid}, {"_id": 0})
    doc["metrics"] = _compute_inventory_metrics(doc)
    return doc


@api.delete("/inventory/{iid}")
async def delete_inventory(iid: str, user=Depends(get_current_user)):
    res = await inventory_col.delete_one({"id": iid, "user_id": user["id"]})
    if not res.deleted_count:
        raise HTTPException(status_code=404, detail="Not found")
    return {"ok": True}


@api.get("/inventory-meta")
async def inventory_meta(user=Depends(get_current_user)):
    return {"sources": SOURCE_TAGS, "platforms": SALE_PLATFORMS, "statuses": STATUSES}


# ---------------- Stats ----------------
@api.get("/stats/home")
async def stats_home(user=Depends(get_current_user)):
    items = await inventory_col.find({"user_id": user["id"]}, {"_id": 0}).to_list(length=5000)
    revenue = 0.0
    profit = 0.0
    total_purchase = 0.0
    inv_value = 0.0
    listed_count = 0
    sold_count = 0
    in_stock_count = 0
    cat_profit: dict = {}
    src_profit: dict = {}
    for it in items:
        st = it.get("status", "in_stock")
        m = _compute_inventory_metrics(it)
        if st == "sold":
            revenue += float(it.get("sale_price") or 0)
            profit += m["net_profit"]
            total_purchase += float(it.get("purchase_price") or 0)
            sold_count += 1
            cat = it.get("category") or "Uncategorized"
            cat_profit[cat] = cat_profit.get(cat, 0) + m["net_profit"]
            src = it.get("source") or "Other"
            src_profit[src] = src_profit.get(src, 0) + m["net_profit"]
        elif st == "listed":
            listed_count += 1
            inv_value += float(it.get("sale_price") or it.get("purchase_price") or 0)
        elif st == "in_stock":
            in_stock_count += 1
            inv_value += float(it.get("purchase_price") or 0)
    roi = round((profit / total_purchase) * 100, 1) if total_purchase > 0 else 0.0
    best_category = max(cat_profit.items(), key=lambda x: x[1])[0] if cat_profit else "—"
    best_source = max(src_profit.items(), key=lambda x: x[1])[0] if src_profit else "—"
    return {
        "total_revenue": round(revenue, 2),
        "total_profit": round(profit, 2),
        "roi_pct": roi,
        "inventory_value": round(inv_value, 2),
        "active_listings": listed_count,
        "items_sold": sold_count,
        "in_stock": in_stock_count,
        "best_category": best_category,
        "best_source": best_source,
    }


@api.get("/stats/sourcing")
async def stats_sourcing(user=Depends(get_current_user)):
    items = await inventory_col.find({"user_id": user["id"]}, {"_id": 0}).to_list(length=5000)
    by_src: dict = {}
    for it in items:
        src = it.get("source") or "Other"
        if src not in by_src:
            by_src[src] = {
                "source": src, "revenue": 0.0, "profit": 0.0, "purchase": 0.0,
                "items": 0, "sold": 0, "days_total": 0, "days_count": 0,
            }
        bucket = by_src[src]
        bucket["items"] += 1
        m = _compute_inventory_metrics(it)
        if it.get("status") == "sold":
            bucket["sold"] += 1
            bucket["revenue"] += float(it.get("sale_price") or 0)
            bucket["profit"] += m["net_profit"]
            bucket["purchase"] += float(it.get("purchase_price") or 0)
            if m["days_to_sell"] is not None:
                bucket["days_total"] += m["days_to_sell"]
                bucket["days_count"] += 1
    rows = []
    for src, b in by_src.items():
        roi = round((b["profit"] / b["purchase"]) * 100, 1) if b["purchase"] > 0 else 0.0
        avg_days = round(b["days_total"] / b["days_count"], 1) if b["days_count"] else None
        rows.append({
            "source": src,
            "revenue": round(b["revenue"], 2),
            "profit": round(b["profit"], 2),
            "roi_pct": roi,
            "items": b["items"],
            "sold": b["sold"],
            "avg_days_to_sell": avg_days,
        })
    rows.sort(key=lambda r: r["profit"], reverse=True)
    best = rows[0]["source"] if rows and rows[0]["profit"] > 0 else "—"
    worst = rows[-1]["source"] if rows and rows[-1]["profit"] < 0 else (rows[-1]["source"] if rows else "—")
    return {"rows": rows, "best_source": best, "worst_source": worst}


@api.get("/stats/reports")
async def stats_reports(period: str = "monthly", user=Depends(get_current_user)):
    """Return time-bucketed profit + revenue + count for a period."""
    items = await inventory_col.find(
        {"user_id": user["id"], "status": "sold"}, {"_id": 0}
    ).to_list(length=5000)
    buckets: dict = {}

    def _key(dt: datetime) -> str:
        if period == "daily":
            return dt.strftime("%Y-%m-%d")
        if period == "weekly":
            iso = dt.isocalendar()
            return f"{iso[0]}-W{iso[1]:02d}"
        if period == "yearly":
            return dt.strftime("%Y")
        return dt.strftime("%Y-%m")

    for it in items:
        try:
            ds = it.get("date_sold") or it.get("created_at")
            dt = datetime.fromisoformat(ds.replace("Z", "+00:00")) if ds else datetime.now(timezone.utc)
        except Exception:
            dt = datetime.now(timezone.utc)
        k = _key(dt)
        if k not in buckets:
            buckets[k] = {"label": k, "revenue": 0.0, "profit": 0.0, "count": 0}
        m = _compute_inventory_metrics(it)
        buckets[k]["revenue"] += float(it.get("sale_price") or 0)
        buckets[k]["profit"] += m["net_profit"]
        buckets[k]["count"] += 1
    rows = sorted(buckets.values(), key=lambda r: r["label"])
    for r in rows:
        r["revenue"] = round(r["revenue"], 2)
        r["profit"] = round(r["profit"], 2)
    totals = {
        "revenue": round(sum(r["revenue"] for r in rows), 2),
        "profit": round(sum(r["profit"] for r in rows), 2),
        "count": sum(r["count"] for r in rows),
    }
    return {"period": period, "rows": rows[-24:], "totals": totals}


# ---------------- Profit Scout Score ----------------
class ScoreIn(BaseModel):
    query: str
    buy_cost: float = 0
    sell_price: float = 0
    ebay_fee_pct: float = 13.25
    shipping_cost: float = 0
    extra_cost: float = 0


@api.post("/score")
async def profit_scout_score(body: ScoreIn, user=Depends(get_current_user)):
    """Compute 0-100 Profit Scout Score from real eBay data + profit math.

    Requires eBay API. If unavailable, returns {"available": false, ...} so the
    UI can render "Live pricing unavailable" instead of fake values.
    """
    if not ebay_configured():
        return {"available": False, "message": "eBay API not configured"}
    try:
        ebay = await ebay_browse_search(body.query)
    except HTTPException as e:
        return {"available": False, "message": e.detail if isinstance(e.detail, str) else "Live pricing unavailable"}
    if not ebay.get("listings"):
        return {"available": False, "message": "No live eBay listings for this query."}

    med = float(ebay.get("median_price") or ebay.get("avg_price") or 0)
    active = int(ebay.get("active_count") or 0)
    sale_price = body.sell_price if body.sell_price > 0 else med
    fee = sale_price * (body.ebay_fee_pct / 100)
    net = sale_price - (body.buy_cost + body.shipping_cost + body.extra_cost + fee)
    roi = (net / body.buy_cost * 100) if body.buy_cost > 0 else 0

    # subscores 0-100 — all from REAL data
    #  - demand proxy: sample_count (unique listings returned)
    demand = min(100, round(int(ebay.get("sample_count") or 0) * 4))
    #  - competition: lower active_count is better (cap at 250 = 0)
    competition = max(0, 100 - min(100, round(active / 2.5)))
    profit_s = max(0, min(100, round(roi)))
    #  - velocity: unknown without sold data — use active-count proxy
    velocity = min(100, max(0, round(100 - (active / 3.0)))) if active else 50
    month = datetime.now().month
    season_table = {1: 50, 2: 50, 3: 55, 4: 60, 5: 65, 6: 70, 7: 70, 8: 65, 9: 65, 10: 75, 11: 90, 12: 95}
    seasonality = season_table.get(month, 60)
    weights = {"demand": 0.25, "competition": 0.2, "profit": 0.3, "velocity": 0.15, "seasonality": 0.1}
    score = round(
        demand * weights["demand"]
        + competition * weights["competition"]
        + profit_s * weights["profit"]
        + velocity * weights["velocity"]
        + seasonality * weights["seasonality"]
    )
    if score >= 70:
        verdict = "BUY"
    elif score >= 45:
        verdict = "MAYBE BUY"
    else:
        verdict = "AVOID"
    return {
        "available": True,
        "score": int(score),
        "verdict": verdict,
        "subscores": {
            "demand": demand,
            "competition": competition,
            "profit": profit_s,
            "velocity": velocity,
            "seasonality": seasonality,
        },
        "net_profit": round(net, 2),
        "roi_pct": round(roi, 1),
        "ebay_snapshot": {
            "active_count": active,
            "avg_price": ebay.get("avg_price"),
            "median_price": ebay.get("median_price"),
            "sample_count": ebay.get("sample_count"),
        },
    }


# ---------------- Settings ----------------
class SettingsIn(BaseModel):
    theme: Optional[str] = None  # light | dark | system
    currency: Optional[str] = None  # USD | EUR | GBP | CAD | AUD
    notifications_enabled: Optional[bool] = None


@api.get("/settings")
async def get_settings(user=Depends(get_current_user)):
    doc = await settings_col.find_one({"user_id": user["id"]}, {"_id": 0}) or {
        "user_id": user["id"],
        "theme": "light",
        "currency": "USD",
        "notifications_enabled": True,
    }
    return doc


@api.post("/settings")
async def update_settings(body: SettingsIn, user=Depends(get_current_user)):
    updates = {k: v for k, v in body.dict().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No fields")
    updates["user_id"] = user["id"]
    await settings_col.update_one({"user_id": user["id"]}, {"$set": updates}, upsert=True)
    doc = await settings_col.find_one({"user_id": user["id"]}, {"_id": 0})
    return doc



# ---------------- Mount ----------------
app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
